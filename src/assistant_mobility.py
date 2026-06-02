import pandas as pd
import numpy as np
import re
import sys
import unicodedata
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

_run_log      = []
_new_ref_entries = {}  # {code: {name, country}} — nouveaux établissements à ajouter à la table

def _log(msg):
    _run_log.append(msg)
    print(msg)

def _strip_accents(s):
    return ''.join(
        c for c in unicodedata.normalize('NFD', str(s).lower())
        if unicodedata.category(c) != 'Mn'
    )

# ---------------------------------------------------------------------------
# Colonnes reconnues
# ---------------------------------------------------------------------------

PARTNER_COLS = [
    "name of institution",
    "host university - current 1. choice",
    "1st choice: host institution",
    "nom université d'origine",
    "nom universite d'origine",
    "host institution",
    "home institution",
]

PROGRAMME_COLS = ["exchange program"]

CODE_COLS = [
    "code institutionnel/inst.no",
    "erasmus code of institution",
    "erasmus code",
    "inst.no",
]

COUNTRY_COLS = [
    "country of home institution",
    "county of home institution",   # typo présent dans certains fichiers BBA
    "pays université d'origine",
    "pays universite d'origine",
]

SEMESTER_COLS = ["semester"]

TEMPLATE_FILE = "Template fichiers suivi IN-OUT.xlsx"

# ---------------------------------------------------------------------------
# Table de zones géographiques
# ---------------------------------------------------------------------------

_ZONE_MAP = {}
for _zone, _countries in [
    ("Europe", [
        # Français
        "allemagne", "autriche", "belgique", "bulgarie", "croatie", "danemark",
        "espagne", "estonie", "finlande", "france", "grece", "hongrie", "irlande",
        "islande", "italie", "lettonie", "lituanie", "luxembourg", "malte",
        "norvege", "pays-bas", "pologne", "portugal", "republique tcheque",
        "roumanie", "royaume-uni", "russie", "slovaquie", "slovenie", "suede",
        "suisse", "turquie", "ukraine", "serbie", "chypre", "georgie", "albanie",
        "bosnie", "macedoine", "montenegro", "liechtenstein", "andorre", "monaco",
        "bielorussie", "moldavie", "kosovo", "tchequie",
        # Anglais
        "germany", "austria", "belgium", "bulgaria", "croatia", "denmark",
        "spain", "estonia", "finland", "greece", "hungary", "ireland", "iceland",
        "italy", "latvia", "lithuania", "netherlands", "norway", "poland",
        "czech republic", "czechia", "romania", "united kingdom", "uk", "russia",
        "slovakia", "slovenia", "sweden", "switzerland", "turkey", "ukraine",
        "serbia", "cyprus", "georgia", "albania", "bosnia", "north macedonia",
        "montenegro", "liechtenstein", "andorra", "monaco", "belarus", "moldova",
        "kosovo", "luxembourg", "malta", "portugal",
    ]),
    ("Ameriques", [
        # Français
        "argentine", "bresil", "canada", "chili", "colombie", "costa rica",
        "etats-unis", "mexique", "perou", "uruguay", "bolivie", "equateur",
        "venezuela", "paraguay", "guatemala", "panama", "cuba", "jamaique",
        "trinite", "honduras", "el salvador", "nicaragua", "republique dominicaine",
        # Anglais
        "argentina", "brazil", "chile", "colombia", "mexico", "peru",
        "united states", "usa", "united states of america", "ecuador", "bolivia",
        "venezuela", "paraguay", "uruguay", "cuba", "jamaica", "trinidad",
        "dominican republic", "guatemala", "panama", "honduras", "el salvador",
        "nicaragua", "costa rica",
    ]),
    ("Asie-Pacifique", [
        # Français
        "australie", "chine", "coree du sud", "coree", "hong kong", "inde",
        "indonesie", "japon", "malaisie", "nouvelle-zelande", "singapour",
        "taiwan", "thailande", "vietnam", "viet nam", "philippines", "mongolie",
        "bangladesh", "sri lanka", "nepal", "pakistan", "myanmar", "cambodge",
        "laos", "nouvelle-caledonie", "fidji", "macao", "kazakhstan", "ouzbekistan",
        "kirghizistan", "tadjikistan", "turkmenistan", "azerbaidjan", "armenie",
        # Anglais
        "australia", "china", "south korea", "korea", "india", "indonesia",
        "japan", "malaysia", "new zealand", "singapore", "thailand", "vietnam",
        "philippines", "mongolia", "sri lanka", "nepal", "pakistan", "myanmar",
        "cambodia", "laos", "fiji", "macau", "kazakhstan", "uzbekistan",
        "kyrgyzstan", "tajikistan", "turkmenistan", "azerbaijan", "armenia",
        "taiwan", "hong kong",
    ]),
    ("Afrique", [
        # Français
        "afrique du sud", "egypte", "maroc", "senegal", "tunisie", "nigeria",
        "ghana", "kenya", "tanzanie", "ethiopie", "cameroun", "cote d'ivoire",
        "zimbabwe", "mozambique", "rwanda", "madagascar", "mali", "burkina faso",
        "algerie", "libye", "soudan", "ouganda", "angola", "zambie", "namibie",
        "botswana", "republique democratique du congo", "niger",
        # Anglais
        "south africa", "egypt", "morocco", "senegal", "tunisia", "nigeria",
        "ghana", "kenya", "tanzania", "ethiopia", "cameroon", "ivory coast",
        "zimbabwe", "mozambique", "rwanda", "madagascar", "mali", "algeria",
        "libya", "sudan", "uganda", "angola", "zambia", "namibia", "botswana",
    ]),
    ("Moyen-Orient", [
        # Français
        "arabie saoudite", "emirats arabes unis", "israel", "jordanie", "koweit",
        "liban", "qatar", "iran", "irak", "bahrain", "oman", "yemen", "syrie",
        # Anglais
        "saudi arabia", "united arab emirates", "uae", "israel", "jordan",
        "kuwait", "lebanon", "iran", "iraq", "bahrain", "oman", "yemen", "syria",
    ]),
]:
    for _c in _countries:
        _ZONE_MAP[_strip_accents(_c)] = _zone

# Mots-clés indiquant que la valeur Country est en réalité un nom d'université → ignorer
_UNI_KEYWORDS = {
    "university", "université", "school", "college", "institute", "business",
    "management", "economics", "polytechnic", "academy", "faculty", "hochschule",
    "universitat", "universidad", "universidade", "universite",
}


# Préfixes Erasmus → pays (ordre du plus long au plus court pour la détection)
_ERASMUS_PREFIX = {
    "SF":  "Finlande",   "FI":  "Finlande",
    "DK":  "Danemark",   "NO":  "Norvège",    "N":   "Norvège",
    "SE":  "Suède",      "S":   "Suède",
    "IS":  "Islande",
    "UK":  "Royaume-Uni","GB":  "Royaume-Uni",
    "IE":  "Irlande",
    "NL":  "Pays-Bas",
    "BE":  "Belgique",   "B":   "Belgique",
    "LU":  "Luxembourg", "L":   "Luxembourg",
    "DE":  "Allemagne",  "D":   "Allemagne",
    "AT":  "Autriche",   "A":   "Autriche",
    "CH":  "Suisse",
    "FR":  "France",     "F":   "France",
    "ES":  "Espagne",    "E":   "Espagne",
    "PT":  "Portugal",   "P":   "Portugal",
    "IT":  "Italie",     "I":   "Italie",
    "GR":  "Grèce",      "EL":  "Grèce",      "G":   "Grèce",
    "TR":  "Turquie",
    "PL":  "Pologne",
    "CZ":  "République tchèque",
    "SK":  "Slovaquie",
    "HU":  "Hongrie",    "H":   "Hongrie",
    "RO":  "Roumanie",
    "BG":  "Bulgarie",
    "HR":  "Croatie",
    "SI":  "Slovénie",
    "EE":  "Estonie",    "LV":  "Lettonie",   "LT":  "Lituanie",
    "CY":  "Chypre",     "MT":  "Malte",
    "RS":  "Serbie",     "MK":  "Macédoine",  "AL":  "Albanie",
    "ME":  "Monténégro", "BA":  "Bosnie",
    "UA":  "Ukraine",    "BY":  "Biélorussie","MD":  "Moldavie",
    "GE":  "Géorgie",    "AM":  "Arménie",    "AZ":  "Azerbaïdjan",
    "RU":  "Russie",
    "IL":  "Israël",     "JO":  "Jordanie",   "LB":  "Liban",
    "TR":  "Turquie",
}

def _country_from_code(code):
    """Dérive le pays depuis le préfixe du code Erasmus si possible."""
    if not code:
        return ""
    code_upper = str(code).strip().upper()
    # Trier par longueur décroissante pour matcher le plus long préfixe d'abord
    for prefix in sorted(_ERASMUS_PREFIX, key=len, reverse=True):
        if code_upper.startswith(prefix):
            return _ERASMUS_PREFIX[prefix]
    return ""


def get_zone(country):
    if pd.isna(country) or not str(country).strip():
        return ""
    return _ZONE_MAP.get(_strip_accents(str(country).strip()), "Autre")


# ---------------------------------------------------------------------------
# Table de référence (codes institutionnels)
# ---------------------------------------------------------------------------

def load_ref_table():
    """Charge la table de référence depuis le template (InstCode → name/country/zone)."""
    tpl = Path(TEMPLATE_FILE)
    if not tpl.exists():
        _log(f"  [WARN] Template '{TEMPLATE_FILE}' introuvable — table de référence ignorée")
        return {}
    try:
        df = pd.read_excel(tpl, sheet_name="Codes institutionnels", engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
        ref = {}
        for _, row in df.iterrows():
            code    = str(row.get("Inst.no",             "") or "").strip().upper()
            name    = str(row.get("Name of institution", "") or "").strip()
            country = str(row.get("Country name",        "") or "").strip()
            # Nettoyer les valeurs parasites issues de runs précédents
            if _strip_accents(code)    in ("nan", "none", ""): code    = ""
            if _strip_accents(name)    in ("nan", "none", ""): name    = ""
            if _strip_accents(country) in ("nan", "none", ""): country = ""
            if code and name:
                ref[code] = {"name": name, "country": country, "zone": get_zone(country)}
        _log(f"  [ref]   Table de référence : {len(ref)} établissements")
        return ref
    except Exception as e:
        _log(f"  [WARN] Erreur lecture table de référence : {e}")
        return {}


# ---------------------------------------------------------------------------
# Fonctions utilitaires
# ---------------------------------------------------------------------------

def norm_partner(name):
    if pd.isna(name):
        return np.nan
    s = str(name).strip().upper()
    s = re.sub(r"\s+", " ", s)
    return s


def extract_year(filename):
    years = re.findall(r"20\d{2}", str(filename))
    return f"{years[0]}-{years[1]}" if len(years) >= 2 else "UNKNOWN"


def extract_programme(filename):
    name = Path(filename).stem
    name = re.sub(r"^(Incoming|Outgoing)\s+", "", name, flags=re.IGNORECASE)
    name = re.sub(r"\s+20\d{2}[-_]20\d{2}\s*$", "", name).strip()
    return _normalise_programme(name)


def _normalise_semester(value):
    """Normalise les intitulés de semestre : supprime l'année, harmonise la casse."""
    if not value or pd.isna(value):
        return None
    s = str(value).strip()
    # Supprimer les années : "2024 / 2025", "2023/2024", "2024-2025", "2024_2025"
    s = re.sub(r'\s*[\-/_]\s*', '/', s)          # normaliser séparateurs
    s = re.sub(r'\s*20\d{2}\s*/\s*20\d{2}', '', s)  # retirer "2024/2025"
    s = re.sub(r'\s*20\d{2}', '', s)              # retirer année seule restante
    s = s.strip(' /-_')
    # Normaliser casse → Title Case
    s = s.title()
    # Mapping des variantes connues → valeur standard
    _map = {
        "Fall Semester":   ["Fall", "Autumn Semester", "Autumn", "Semestre D'Automne"],
        "Spring Semester": ["Spring", "Semestre De Printemps"],
        "Whole Year":      ["Whole Year", "Whole Year", "Full Year", "Annual", "Year"],
        "Summer School":   ["Summer School", "Ss", "Ete"],
        "Winter School":   ["Winter School", "Ws", "Hiver"],
        "Summer Semester": ["Summer Semester", "Summer"],
    }
    s_norm = s.strip()
    for standard, variants in _map.items():
        if s_norm == standard:
            return standard
        for v in variants:
            if v.lower() == s_norm.lower():
                return standard
    return s_norm if s_norm else None


def _normalise_programme(value):
    """Applique les alias et normalise la casse."""
    if not value or pd.isna(value):
        return value
    s = str(value).strip()
    for pattern, replacement in [
        (r"^MIM\s+DD$", "PGE DD"),
        (r"^MIM$",      "PGE"),
    ]:
        if re.match(pattern, s, re.IGNORECASE):
            return replacement
    return s


def find_student_list_header(xls, sh):
    df_raw = pd.read_excel(xls, sh, header=None, engine="openpyxl", nrows=20)
    for i, row in df_raw.iterrows():
        row_str = " | ".join(str(v).lower().strip() for v in row if pd.notna(v))
        for pc in PARTNER_COLS:
            if pc in row_str:
                return i, pc
    return None, None


def _year_in_sheet(sh_name, file_year):
    years = re.findall(r"20\d{2}", file_year)
    sh_lower = sh_name.lower()
    short_years = [y[-2:] for y in years]
    return all(sy in sh_lower for sy in short_years)


def extract_from_multiline(text):
    if not isinstance(text, str) or not text.upper().startswith("ID ("):
        return text
    match = re.search(
        r"1st choice: host institution:\s*(.+?)(\n|_x000D_|$)", text, re.IGNORECASE
    )
    return match.group(1).strip() if match else None


def read_partner_row_data(ws, header_row, matched_col):
    """Lit partner + InstCode + Country + Semester en ignorant les lignes barrées."""
    openpyxl_header = header_row + 1
    partner_col_idx = code_col_idx = country_col_idx = semester_col_idx = programme_col_idx = None

    for cell in ws[openpyxl_header]:
        if cell.value is None:
            continue
        val = str(cell.value).lower().strip()
        if matched_col in val and partner_col_idx is None:
            partner_col_idx = cell.column
        if code_col_idx is None:
            for c in CODE_COLS:
                if c in val:
                    code_col_idx = cell.column
                    break
        if country_col_idx is None:
            for c in COUNTRY_COLS:
                if c in val:
                    country_col_idx = cell.column
                    break
        if semester_col_idx is None and any(c in val for c in SEMESTER_COLS):
            semester_col_idx = cell.column
        if programme_col_idx is None and any(c in val for c in PROGRAMME_COLS):
            programme_col_idx = cell.column

    if partner_col_idx is None:
        return pd.DataFrame()

    def _get(row, idx):
        if idx is None:
            return None
        v = row[idx - 1].value
        return str(v).strip() if v is not None else None

    rows = []
    for row in ws.iter_rows(min_row=openpyxl_header + 1):
        cell = row[partner_col_idx - 1]
        if cell.value is None or str(cell.value).strip() == "":
            continue
        if cell.font and cell.font.strike:
            continue
        rows.append({
            "partner_raw":    str(cell.value),
            "InstCode":       _get(row, code_col_idx),
            "Country":        _get(row, country_col_idx),
            "Semester":       _get(row, semester_col_idx),
            "ProgrammeRaw":   _get(row, programme_col_idx),
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame()


def find_tcd_header(xls, sh):
    df_raw = pd.read_excel(xls, sh, header=None, engine="openpyxl", nrows=30)
    for i, row in df_raw.iterrows():
        row_str = " ".join(str(v).lower() for v in row if pd.notna(v))
        if "tiquettes" in row_str:
            return i
    return None


# ---------------------------------------------------------------------------
# Chargement d'un fichier
# ---------------------------------------------------------------------------

def load_file(filepath, flux):
    try:
        xls = pd.ExcelFile(filepath, engine="openpyxl")
        wb  = load_workbook(filepath, data_only=True)
    except Exception:
        _log(f"  [ERREUR] Impossible de lire : {Path(filepath).name}")
        return pd.DataFrame()

    year      = extract_year(str(filepath))
    programme = extract_programme(str(filepath))
    count_col = "Incoming" if flux == "incoming" else "Outgoing"

    best_df = best_sh = None
    best_total = 0
    year_df = year_sh = None
    year_total = 0

    for sh in xls.sheet_names:
        header_row, matched_col = find_student_list_header(xls, sh)
        if header_row is None:
            continue

        ws = wb[sh] if sh in wb.sheetnames else None
        if ws is not None:
            row_data = read_partner_row_data(ws, header_row, matched_col)
        else:
            df = pd.read_excel(xls, sh, header=header_row, engine="openpyxl")
            cols_lower = {str(c).lower().strip(): c for c in df.columns}
            partner_col = cols_lower.get(matched_col)
            if partner_col is None:
                continue
            row_data = pd.DataFrame({"partner_raw": df[partner_col].dropna().astype(str)})

        if row_data.empty:
            continue

        row_data["partner_raw"] = row_data["partner_raw"].apply(extract_from_multiline)
        row_data = row_data.dropna(subset=["partner_raw"])
        row_data = row_data[row_data["partner_raw"].str.strip() != ""]
        if len(row_data) < 3:
            continue

        row_data["Partner"] = row_data["partner_raw"].apply(norm_partner)
        row_data = row_data.dropna(subset=["Partner"])

        total = len(row_data)
        if total > best_total:
            best_total = total
            best_df    = row_data.copy()
            best_sh    = sh
        if _year_in_sheet(sh, year) and total > year_total:
            year_total = total
            year_df    = row_data.copy()
            year_sh    = sh

    result     = year_df    if year_df    is not None else best_df
    result_sh  = year_sh    if year_df    is not None else best_sh
    result_tot = year_total if year_df    is not None else best_total

    if result is not None:
        result["AcademicYear"] = year
        # Exchange program du fichier prime sur le nom de fichier
        if "ProgrammeRaw" in result.columns and result["ProgrammeRaw"].notna().any():
            result["Programme"] = result["ProgrammeRaw"].apply(_normalise_programme).fillna(programme)
        else:
            result["Programme"] = programme
        result[count_col] = 1
        _log(f"  [liste] {Path(filepath).name:45s} onglet '{result_sh}' -> {result_tot} etudiants")
        wb.close()
        for c in ["InstCode", "Country", "Semester"]:
            if c not in result.columns:
                result[c] = None
        result["Semester"] = result["Semester"].apply(_normalise_semester)
        # SS et WS : programme = semestre, on auto-remplit si vide
        ss_ws = {"Summer School", "Winter School"}
        if programme in ss_ws:
            result["Semester"] = result["Semester"].fillna(programme).replace("", programme)
        return result[["Partner", "InstCode", "Country", "Semester",
                        "AcademicYear", "Programme", count_col]]

    # --- Priorité 2 : TCD ---
    for sh in xls.sheet_names:
        header_row = find_tcd_header(xls, sh)
        if header_row is None:
            continue
        df = pd.read_excel(xls, sh, header=header_row, engine="openpyxl")
        first_col, value_col = df.columns[0], df.columns[1]
        rows = [
            {"Partner":  norm_partner(str(row[first_col])),
             count_col:  int(row[value_col]),
             "InstCode": None, "Country": None, "Semester": None}
            for _, row in df.iterrows()
            if isinstance(row[first_col], str)
            and isinstance(row[value_col], (int, float))
            and "TOTAL" not in str(row[first_col]).upper()
            and str(row[first_col]).strip() != ""
        ]
        if rows:
            tdf = pd.DataFrame(rows)
            tdf["AcademicYear"] = year
            tdf["Programme"]    = programme
            _log(f"  [TCD]   {Path(filepath).name:45s} onglet '{sh}' -> {tdf[count_col].sum()} etudiants")
            wb.close()
            return tdf[["Partner", "InstCode", "Country", "Semester",
                         "AcademicYear", "Programme", count_col]]

    wb.close()
    _log(f"  [SKIP]  {Path(filepath).name} : aucun format reconnu")
    return pd.DataFrame()


# ---------------------------------------------------------------------------
# Enrichissement via table de référence
# ---------------------------------------------------------------------------

def _apply_ref_to_df(df, ref):
    """Normalise Partner depuis InstCode, enrichit Country et Zone.
    Enregistre dans _new_ref_entries les codes inconnus trouvés dans les fichiers."""
    if df.empty:
        df["Zone"] = pd.Series(dtype=str)
        return df

    def _clean_country(val):
        """Retourne None si la valeur est vide, 'nan', ou ressemble à un nom d'université."""
        s = str(val or "").strip()
        if not s or _strip_accents(s) in ("nan", "none", "n/a", ""):
            return ""
        if any(kw in s.lower() for kw in _UNI_KEYWORDS):
            return ""
        return s

    def resolve(row):
        code    = str(row.get("InstCode") or "").strip().upper()
        country = _clean_country(row.get("Country"))
        if code and code in ref:
            entry       = ref[code]
            name        = entry["name"]
            ref_country = entry["country"] if _strip_accents(entry["country"]) not in ("nan","none","") else ""
            country     = country or ref_country or _country_from_code(code)
            return pd.Series({"Partner": name, "InstCode": code,
                               "Country": country, "Zone": get_zone(country)})
        if code and code not in ("NAN", ""):
            # Code présent dans le fichier mais absent de la table → noter pour ajout
            if code not in _new_ref_entries:
                _new_ref_entries[code] = {
                    "name":    str(row.get("Partner") or "").strip(),
                    "country": country,
                }
        clean_code = code if code not in ("NAN", "NONE", "") else None
        return pd.Series({"Partner": row["Partner"], "InstCode": clean_code,
                           "Country": country, "Zone": get_zone(country)})

    extra = df.apply(resolve, axis=1)
    df = df.copy()
    df["Partner"]  = extra["Partner"]
    df["InstCode"] = extra["InstCode"]
    df["Country"]  = extra["Country"]
    df["Zone"]     = extra["Zone"]
    return df


def _update_ref_table(new_entries):
    """Ajoute les nouveaux établissements à la feuille Codes institutionnels du template.
    N'efface jamais les entrées existantes."""
    if not new_entries:
        return
    tpl = Path(TEMPLATE_FILE)
    if not tpl.exists():
        _log(f"  [WARN] Template introuvable — impossible d'ajouter {len(new_entries)} entrée(s)")
        return
    try:
        wb = load_workbook(tpl)
        ws = wb["Codes institutionnels"]
        last_row = ws.max_row
        added = 0
        for code, entry in sorted(new_entries.items()):
            last_row += 1
            ws.cell(last_row, 1, entry["country"])
            ws.cell(last_row, 2, entry["name"])
            ws.cell(last_row, 3, code)
            added += 1
        wb.save(tpl)
        _log(f"  [ref]   {added} nouvel(s) établissement(s) ajouté(s) à la table de référence")
        for code, entry in sorted(new_entries.items()):
            _log(f"           + {code} : {entry['name']} ({entry['country'] or 'pays inconnu'})")
    except PermissionError:
        _log(f"  [WARN] Template ouvert dans Excel — nouveaux codes non sauvegardés : {list(new_entries.keys())}")
    except Exception as e:
        _log(f"  [WARN] Erreur mise à jour table de référence : {e}")


# ---------------------------------------------------------------------------
# Formatage Excel
# ---------------------------------------------------------------------------

def apply_table_format(ws, table_name, balance_col_idx=None):
    if ws.max_row < 2:
        return
    data_last_row = ws.max_row
    headers       = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    numeric_cols  = {"Incoming", "Outgoing", "Balance"}

    ws.freeze_panes = "A2"

    max_col_letter = get_column_letter(ws.max_column)
    ref = f"A1:{max_col_letter}{data_last_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True,
        showFirstColumn=False, showLastColumn=False, showColumnStripes=False,
    )
    ws.add_table(table)

    totals_row       = data_last_row + 1
    first_label_done = False
    for c, header in enumerate(headers, 1):
        cell = ws.cell(totals_row, c)
        if header in numeric_cols:
            col_letter = get_column_letter(c)
            cell.value = f"=SUBTOTAL(9,{col_letter}2:{col_letter}{data_last_row})"
            cell.font  = Font(bold=True)
        elif not first_label_done:
            cell.value       = "TOTAL"
            cell.font        = Font(bold=True)
            first_label_done = True

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)

    # Couleur en-tête Incoming (bleu foncé) et Outgoing (gris)
    outgoing_col_idx = None
    for c, header in enumerate(headers, 1):
        cell = ws.cell(1, c)
        if header == "Incoming":
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(bold=True, color="FFFFFF")
        elif header == "Outgoing":
            cell.fill = PatternFill("solid", fgColor="808080")
            cell.font = Font(bold=True, color="FFFFFF")
            outgoing_col_idx = c

    # Alternance gris clair / gris foncé sur la colonne Outgoing
    if outgoing_col_idx:
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=data_last_row)):
            cell = row[outgoing_col_idx - 1]
            cell.fill = PatternFill("solid", fgColor="D9D9D9" if i % 2 == 0 else "BFBFBF")

    if balance_col_idx:
        for row in ws.iter_rows(min_row=2, max_row=data_last_row):
            cell = row[balance_col_idx - 1]
            if isinstance(cell.value, (int, float)):
                v = cell.value
                if v > 0:
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")   # vert
                elif -3 <= v <= -1:
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")   # jaune
                elif -6 <= v <= -4:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")   # orange
                elif v < -6:
                    cell.fill = PatternFill("solid", fgColor="FF6666")   # rouge atténué


# ---------------------------------------------------------------------------
# Logs
# ---------------------------------------------------------------------------



def _parse_log_entry(entry):
    entry      = entry.strip()
    status_map = {"liste": "Liste brute", "TCD": "TCD", "SKIP": "Non reconnu", "ERREUR": "Erreur"}
    sm         = re.match(r'\[(liste|TCD|SKIP|ERREUR)\]', entry)
    status     = status_map.get(sm.group(1), "?") if sm else "?"
    detail     = re.search(
        r'\[(?:liste|TCD|SKIP|ERREUR)\]\s+(.+?\.xlsx\s*)\s+onglet \'(.+?)\' -> (\d+)', entry)
    if detail:
        return status, detail.group(1).strip(), detail.group(2), int(detail.group(3))
    fm = re.search(r'\[(?:liste|TCD|SKIP|ERREUR)\]\s+(.+?\.xlsx)', entry)
    return status, fm.group(1).strip() if fm else entry, "-", 0


def _write_infos_sheet(wb, run_time, summary):
    from openpyxl.styles import Alignment, Border, Side
    ws = wb.create_sheet("Infos", 0)

    C_TITLE_BG  = "1F4E79"
    C_HEADER_BG = "2E75B6"
    C_SECTION   = "D6E4F0"
    C_ROW_ALT   = "F2F7FC"
    C_WARN_BG   = "FDECEA"
    C_WARN_TXT  = "C00000"

    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(r, c, value="", bold=False, color=None, bg=None, align="left", size=11):
        cl = ws.cell(r, c, value)
        cl.font      = Font(bold=bold, color=color or "000000", size=size, name="Calibri")
        if bg:
            cl.fill  = PatternFill("solid", fgColor=bg)
        cl.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        return cl

    ws.merge_cells("A1:D1")
    cell(1, 1, "Balance des echanges internationaux",
         bold=True, color="FFFFFF", bg=C_TITLE_BG, align="center", size=14)
    ws.row_dimensions[1].height = 30

    meta = [
        ("Date de generation",    run_time.strftime("%d/%m/%Y  %H:%M:%S")),
        ("Incoming total",        int(summary["Incoming"].sum())),
        ("Outgoing total",        int(summary["Outgoing"].sum())),
        ("Partenaires distincts", int(summary["Partner"].nunique())),
        ("Annees couvertes",      ", ".join(sorted(summary["AcademicYear"].dropna().unique()))),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        cell(i, 1, label, bold=True)
        cell(i, 2, value)
    ws.row_dimensions[2].height = 6

    row = 10
    ws.merge_cells(f"A{row}:D{row}")
    cell(row, 1, "Fichiers traites", bold=True, bg=C_SECTION)
    ws.row_dimensions[row].height = 18
    row += 1

    for c, h in enumerate(["Fichier", "Onglet", "Etudiants", "Source"], 1):
        cl = cell(row, c, h, bold=True, color="FFFFFF", bg=C_HEADER_BG, align="center")
        cl.border = border
    ws.row_dimensions[row].height = 18
    row += 1

    ok_entries = [e for e in _run_log if "[liste]" in e or "[TCD]" in e]
    for i, entry in enumerate(ok_entries):
        status, fname, sheet, count = _parse_log_entry(entry)
        bg = C_ROW_ALT if i % 2 == 0 else None
        for c, val in enumerate([fname, sheet, count, status], 1):
            cl = cell(row, c, val, bg=bg, align="center" if c in (3, 4) else "left")
            cl.border = border
        row += 1

    warnings = [e for e in _run_log if "[SKIP]" in e or "[ERREUR]" in e]
    if warnings:
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        cell(row, 1, "Fichiers non pris en compte", bold=True, color=C_WARN_TXT, bg="FDECEA")
        ws.row_dimensions[row].height = 18
        row += 1
        for entry in warnings:
            _, fname, _, _ = _parse_log_entry(entry)
            ws.merge_cells(f"A{row}:D{row}")
            cell(row, 1, fname, color=C_WARN_TXT, bg=C_WARN_BG)
            row += 1
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        cell(row, 1,
             "Verifiez que le titre respecte : Incoming [Programme] [Annee].xlsx",
             color=C_WARN_TXT)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14


# ---------------------------------------------------------------------------
# Construction de la balance
# ---------------------------------------------------------------------------

def build_global_balance(incoming_files, outgoing_files, output="Balance_Echanges.xlsx"):
    _run_log.clear()
    _new_ref_entries.clear()
    run_time = datetime.now()

    print("=== Chargement table de reference ===")
    ref = load_ref_table()

    print("\n=== Lecture INCOMING ===")
    inc_parts = [load_file(f, "incoming") for f in incoming_files if Path(f).exists()]

    print("\n=== Lecture OUTGOING ===")
    out_parts = [load_file(f, "outgoing") for f in outgoing_files if Path(f).exists()]

    inc_frames = [d for d in inc_parts if not d.empty]
    out_frames = [d for d in out_parts if not d.empty]

    _EMPTY_INC = pd.DataFrame(columns=["Partner","InstCode","Country","Semester",
                                        "AcademicYear","Programme","Incoming"])
    _EMPTY_OUT = pd.DataFrame(columns=["Partner","InstCode","Country","Semester",
                                        "AcademicYear","Programme","Outgoing"])

    incoming_all = pd.concat(inc_frames, ignore_index=True) if inc_frames else _EMPTY_INC
    outgoing_all = pd.concat(out_frames, ignore_index=True) if out_frames else _EMPTY_OUT

    incoming_all = _apply_ref_to_df(incoming_all, ref)
    outgoing_all = _apply_ref_to_df(outgoing_all, ref)

    _PARTNER_EXCLUDE = re.compile(
        r"free\s*mover|study\s*abroad|em\s*lyon", re.IGNORECASE
    )
    for df in [incoming_all, outgoing_all]:
        mask = df["Partner"].astype(str).str.contains(_PARTNER_EXCLUDE, na=False)
        df.drop(df[mask].index, inplace=True)

    # Table de métadonnées : une seule ligne par Partner (InstCode non-null prioritaire)
    meta_all = pd.concat([
        incoming_all[["Partner", "InstCode", "Country", "Zone"]],
        outgoing_all[["Partner", "InstCode", "Country", "Zone"]],
    ], ignore_index=True).dropna(subset=["Partner"])
    meta_all = (
        meta_all
        .sort_values("InstCode", na_position="last")
        .drop_duplicates(subset=["Partner"])
        [["Partner", "InstCode", "Country", "Zone"]]
    )

    # --- Agrégation détaillée (par programme + semestre uniquement) ---
    detail_key = ["AcademicYear", "Programme", "Partner", "Semester"]
    incoming_all["Semester"] = incoming_all["Semester"].astype(object)
    outgoing_all["Semester"] = outgoing_all["Semester"].astype(object)
    inc_agg = (incoming_all.groupby(detail_key, as_index=False, dropna=False)["Incoming"].sum()
               if not incoming_all.empty else pd.DataFrame(columns=detail_key + ["Incoming"]))
    out_agg = (outgoing_all.groupby(detail_key, as_index=False, dropna=False)["Outgoing"].sum()
               if not outgoing_all.empty else pd.DataFrame(columns=detail_key + ["Outgoing"]))

    inc_agg["Semester"] = inc_agg["Semester"].astype(object)
    out_agg["Semester"] = out_agg["Semester"].astype(object)

    detail = pd.merge(inc_agg, out_agg, on=detail_key, how="outer")
    detail["Incoming"] = detail["Incoming"].fillna(0).astype(int)
    detail["Outgoing"] = detail["Outgoing"].fillna(0).astype(int)
    detail["Balance"]  = detail["Incoming"] - detail["Outgoing"]
    detail = detail.merge(meta_all, on="Partner", how="left")
    detail = detail.sort_values(["AcademicYear", "Partner", "Programme"]).reset_index(drop=True)
    detail = detail[["AcademicYear", "Programme", "Partner", "InstCode",
                     "Country", "Zone", "Semester", "Incoming", "Outgoing", "Balance"]]
    detail = detail.rename(columns={"Programme": "Exchange_program"})

    # --- Résumé (tous programmes confondus) ---
    summary = detail.groupby(["AcademicYear", "Partner"], as_index=False).agg(
        Incoming=("Incoming", "sum"),
        Outgoing=("Outgoing", "sum"),
    )
    summary["Balance"] = summary["Incoming"] - summary["Outgoing"]
    summary = summary.merge(meta_all, on="Partner", how="left")
    summary = summary[["AcademicYear", "Partner", "InstCode",
                        "Country", "Zone", "Incoming", "Outgoing", "Balance"]]

    # Liste cumulative : conserver les partenaires de toutes les années (0 si absent)
    all_years    = sorted(summary["AcademicYear"].dropna().unique())
    all_partners = meta_all[["Partner","InstCode","Country","Zone"]].drop_duplicates(subset=["Partner"])

    # Ajouter les universités du template absentes cette année (avec 0)
    if ref:
        ref_partners = pd.DataFrame([
            {"Partner": v["name"], "InstCode": k, "Country": v["country"], "Zone": v["zone"]}
            for k, v in ref.items() if v["name"]
        ])
        existing_names = set(all_partners["Partner"].str.strip().str.lower())
        ref_new = ref_partners[~ref_partners["Partner"].str.strip().str.lower().isin(existing_names)]
        all_partners = pd.concat([all_partners, ref_new], ignore_index=True)

    if all_years:
        cross = pd.DataFrame({"AcademicYear": all_years}).merge(all_partners, how="cross")
        summary = cross.merge(
            summary[["AcademicYear","Partner","Incoming","Outgoing","Balance"]],
            on=["AcademicYear","Partner"], how="left"
        )
        summary[["Incoming","Outgoing","Balance"]] = (
            summary[["Incoming","Outgoing","Balance"]].fillna(0).astype(int)
        )
        summary = summary.sort_values(["AcademicYear","Partner"]).reset_index(drop=True)

    # --- Écriture Excel ---
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="Resume",        index=False)
            detail.to_excel(writer,  sheet_name="Par_Programme", index=False)
    except PermissionError:
        print(f"\n[ERREUR FICHIER OUVERT] {output} est ouvert dans Excel.")
        print("Fermez-le puis relancez.")

        sys.exit(1)

    wb    = load_workbook(output)
    ws_r  = wb["Resume"]
    hdr_r = [c.value for c in ws_r[1]]
    apply_table_format(ws_r, "TableResume",
                       hdr_r.index("Balance") + 1 if "Balance" in hdr_r else None)

    ws_d  = wb["Par_Programme"]
    hdr_d = [c.value for c in ws_d[1]]
    apply_table_format(ws_d, "TableDetail",
                       hdr_d.index("Balance") + 1 if "Balance" in hdr_d else None)

    _write_infos_sheet(wb, run_time, summary)
    wb.save(output)

    # Ajouter les nouveaux établissements à la table de référence
    _update_ref_table(_new_ref_entries)

    print(f"\n=== RESULTAT ===")
    print(f"[OK] {output}")
    print(f"     Incoming total : {summary['Incoming'].sum()}")
    print(f"     Outgoing total : {summary['Outgoing'].sum()}")
    print(f"     Partenaires    : {summary['Partner'].nunique()}")
    print(f"     Annees         : {all_years}")


if __name__ == "__main__":
    incoming_files = [str(f) for f in Path(".").glob("Incoming*.xlsx")]
    outgoing_files = [str(f) for f in Path(".").glob("Outgoing*.xlsx")]
    build_global_balance(incoming_files, outgoing_files)
